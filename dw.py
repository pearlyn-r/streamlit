import pandas as pd
import numpy as np
from openpyxl import load_workbook

def process_excel_file(file_path, k_col_index, l_col_index, i_col_index, j_col_index):
    """
    Process Excel file by combining sheets with additional columns using openpyxl.
    
    Parameters:
    - file_path: path to Excel file
    - k_col_index: column index for final_rating (0-based, e.g., 0 for column A)
    - l_col_index: column index for final_rating (0-based, e.g., 1 for column B)
    - i_col_index: column index for final_product (0-based)
    - j_col_index: column index for final_product (0-based)
    """
    
    # Load workbook using openpyxl
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    
    print(f"Found {len(sheet_names)} sheets: {sheet_names}")
    
    # Skip first 3 sheets and process the rest
    sheets_to_process = sheet_names[3:]
    print(f"Processing sheets: {sheets_to_process}")
    
    combined_dfs = []
    
    for sheet_name in sheets_to_process:
        # Get the worksheet
        worksheet = workbook[sheet_name]
        
        # Convert worksheet to DataFrame
        df = worksheet_to_dataframe(worksheet)
        
        if df.empty:
            print(f"Warning: Sheet '{sheet_name}' is empty, skipping...")
            continue
        
        # Add date column with sheet name
        df['date'] = sheet_name
        
        # Get column names for the specified indices
        col_names = df.columns.tolist()
        
        # Create final_rating column using if-else logic
        if k_col_index < len(col_names) and l_col_index < len(col_names):
            k_col_name = col_names[k_col_index]
            l_col_name = col_names[l_col_index]
            
            # Apply if-else logic: if k_col is null/blank, use l_col, otherwise use k_col
            def get_final_rating(row):
                k_value = row[k_col_name]
                l_value = row[l_col_name]
                
                # Check if k_value is null, blank, or empty string
                if pd.isna(k_value) or k_value == '' or k_value == ' ' or str(k_value).strip() == '':
                    return l_value
                else:
                    return k_value
            
            df['final_rating'] = df.apply(get_final_rating, axis=1)
        else:
            print(f"Warning: Column indices {k_col_index} or {l_col_index} out of range for sheet {sheet_name}")
            df['final_rating'] = None
        
        # Create final_product column using if-else logic
        if i_col_index < len(col_names) and j_col_index < len(col_names):
            i_col_name = col_names[i_col_index]
            j_col_name = col_names[j_col_index]
            
            # Apply if-else logic: if i_col is null/blank, use j_col, otherwise use i_col
            def get_final_product(row):
                i_value = row[i_col_name]
                j_value = row[j_col_name]
                
                # Check if i_value is null, blank, or empty string
                if pd.isna(i_value) or i_value == '' or i_value == ' ' or str(i_value).strip() == '':
                    return j_value
                else:
                    return i_value
            
            df['final_product'] = df.apply(get_final_product, axis=1)
        else:
            print(f"Warning: Column indices {i_col_index} or {j_col_index} out of range for sheet {sheet_name}")
            df['final_product'] = None
        
        combined_dfs.append(df)
        print(f"Processed sheet '{sheet_name}' with {len(df)} rows")
    
    # Close the workbook
    workbook.close()
    
    # Combine all DataFrames
    if combined_dfs:
        final_df = pd.concat(combined_dfs, ignore_index=True)
        print(f"Combined DataFrame created with {len(final_df)} total rows")
        return final_df
    else:
        print("No sheets to process")
        return pd.DataFrame()

def worksheet_to_dataframe(worksheet):
    """
    Convert openpyxl worksheet to pandas DataFrame.
    """
    # Get all data from worksheet
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(row)
    
    if not data:
        return pd.DataFrame()
    
    # First row as headers
    headers = data[0]
    
    # Clean headers - remove None values and convert to string
    cleaned_headers = []
    for i, header in enumerate(headers):
        if header is None:
            cleaned_headers.append(f"Column_{i}")
        else:
            cleaned_headers.append(str(header))
    
    # Rest of the data
    rows_data = data[1:]
    
    # Create DataFrame
    df = pd.DataFrame(rows_data, columns=cleaned_headers)
    
    # Remove completely empty rows
    df = df.dropna(how='all')
    
    return df

def get_column_by_index(df, col_index):
    """
    Safely get column by index.
    """
    if col_index < len(df.columns):
        return df.iloc[:, col_index]
    else:
        return pd.Series([None] * len(df))

# Alternative version with column names instead of indices
def process_excel_file_by_names(file_path, k_col_name, l_col_name, i_col_name, j_col_name):
    """
    Alternative version that uses column names instead of indices.
    """
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames[3:]  # Skip first 3 sheets
    
    combined_dfs = []
    
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        df = worksheet_to_dataframe(worksheet)
        
        if df.empty:
            continue
            
        df['date'] = sheet_name
        
        # Create final_rating column using if-else logic
        if k_col_name in df.columns and l_col_name in df.columns:
            def get_final_rating(row):
                k_value = row[k_col_name]
                l_value = row[l_col_name]
                
                # Check if k_value is null, blank, or empty string
                if pd.isna(k_value) or k_value == '' or k_value == ' ' or str(k_value).strip() == '':
                    return l_value
                else:
                    return k_value
            
            df['final_rating'] = df.apply(get_final_rating, axis=1)
        else:
            df['final_rating'] = None
            
        # Create final_product column using if-else logic
        if i_col_name in df.columns and j_col_name in df.columns:
            def get_final_product(row):
                i_value = row[i_col_name]
                j_value = row[j_col_name]
                
                # Check if i_value is null, blank, or empty string
                if pd.isna(i_value) or i_value == '' or i_value == ' ' or str(i_value).strip() == '':
                    return j_value
                else:
                    return i_value
            
            df['final_product'] = df.apply(get_final_product, axis=1)
        else:
            df['final_product'] = None
            
        combined_dfs.append(df)
    
    workbook.close()
    return pd.concat(combined_dfs, ignore_index=True) if combined_dfs else pd.DataFrame()

# Example usage
if __name__ == "__main__":
    # CUSTOMIZE THESE PARAMETERS:
    file_path = "your_excel_file.xlsx"  # Replace with your Excel file path
    
    # Column indices (0-based). For example:
    # Column A = 0, Column B = 1, Column C = 2, etc.
    k_col_index = 2  # Replace with actual column index for final_rating
    l_col_index = 3  # Replace with actual column index for final_rating
    i_col_index = 4  # Replace with actual column index for final_product
    j_col_index = 5  # Replace with actual column index for final_product
    
    try:
        # Process the Excel file
        result_df = process_excel_file(file_path, k_col_index, l_col_index, i_col_index, j_col_index)
        
        # Display basic info about the result
        print("\nFinal DataFrame Info:")
        print(f"Shape: {result_df.shape}")
        print(f"Columns: {list(result_df.columns)}")
        
        # Display first few rows
        print("\nFirst 5 rows:")
        print(result_df.head())
        
        # Save to new Excel file using openpyxl
        output_file = "combined_data.xlsx"
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='Combined_Data', index=False)
        print(f"\nCombined data saved to: {output_file}")
        
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found")
    except Exception as e:
        print(f"Error processing file: {str(e)}")

# Additional utility function to inspect Excel file structure
def inspect_excel_file(file_path):
    """
    Utility function to inspect Excel file structure.
    """
    workbook = load_workbook(file_path, read_only=True)
    
    print(f"Excel file: {file_path}")
    print(f"Total sheets: {len(workbook.sheetnames)}")
    print("Sheet names:")
    
    for i, sheet_name in enumerate(workbook.sheetnames):
        worksheet = workbook[sheet_name]
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        print(f"  {i+1}. '{sheet_name}' - {max_row} rows, {max_col} columns")
        
        # Show first row (headers) if exists
        if max_row > 0:
            first_row = [cell.value for cell in worksheet[1]]
            print(f"     Headers: {first_row[:10]}...")  # Show first 10 headers
    
    workbook.close()
