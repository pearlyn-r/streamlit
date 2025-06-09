import pandas as pd
import numpy as np
from openpyxl import load_workbook

def process_excel_file(file_path, k_col_index, l_col_index, i_col_index, j_col_index):
    """
    Process Excel file by combining sheets and creating 'final_rating' and 'final_product'
    using efficient vectorized logic.
    """
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames[3:]  # Skip first 3 sheets
    
    print(f"Processing {len(sheet_names)} sheets: {sheet_names}")
    combined_dfs = []
    
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        df = worksheet_to_dataframe(worksheet)
        
        if df.empty:
            print(f"Skipping empty sheet: {sheet_name}")
            continue
        
        df['date'] = sheet_name
        col_names = df.columns.tolist()

        # Validate and assign final_rating
        if k_col_index < len(col_names) and l_col_index < len(col_names):
            k_col, l_col = col_names[k_col_index], col_names[l_col_index]
            df['final_rating'] = df[k_col].where(df[k_col].notna() & (df[k_col].astype(str).str.strip() != ''), df[l_col])
        else:
            print(f"Invalid indices for final_rating: {k_col_index}, {l_col_index}")
            df['final_rating'] = None
        
        # Validate and assign final_product
        if i_col_index < len(col_names) and j_col_index < len(col_names):
            i_col, j_col = col_names[i_col_index], col_names[j_col_index]
            df['final_product'] = df[i_col].where(df[i_col].notna() & (df[i_col].astype(str).str.strip() != ''), df[j_col])
        else:
            print(f"Invalid indices for final_product: {i_col_index}, {j_col_index}")
            df['final_product'] = None
        
        combined_dfs.append(df)
        print(f"Processed sheet: {sheet_name}, rows: {len(df)}")
    
    workbook.close()

    if combined_dfs:
        final_df = pd.concat(combined_dfs, ignore_index=True)
        print(f"Combined {len(final_df)} rows from all sheets")
        return final_df
    else:
        print("No valid data to combine.")
        return pd.DataFrame()


def worksheet_to_dataframe(worksheet):
    """
    Convert an openpyxl worksheet to a pandas DataFrame.
    """
    data = [row for row in worksheet.iter_rows(values_only=True)]
    
    if not data:
        return pd.DataFrame()

    headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(data[0])]
    df = pd.DataFrame(data[1:], columns=headers)
    df.dropna(how='all', inplace=True)
    return df


def inspect_excel_file(file_path):
    """
    Utility to print structure of Excel file.
    """
    workbook = load_workbook(file_path, read_only=True)
    print(f"Inspecting: {file_path}")
    for i, sheet in enumerate(workbook.sheetnames):
        ws = workbook[sheet]
        print(f"Sheet {i+1}: '{sheet}' - {ws.max_row} rows x {ws.max_column} cols")
        headers = [cell.value for cell in ws[1]]
        print(f"  Headers: {headers[:10]}...")
    workbook.close()


# Example usage
if __name__ == "__main__":
    file_path = "your_excel_file.xlsx"  # Update this
    k_col_index = 2
    l_col_index = 3
    i_col_index = 4
    j_col_index = 5
    
    try:
        df_result = process_excel_file(file_path, k_col_index, l_col_index, i_col_index, j_col_index)
        print("\nResulting DataFrame:")
        print(df_result.head())

        if not df_result.empty:
            df_result.to_excel("combined_data.xlsx", index=False)
            print("Saved combined data to 'combined_data.xlsx'")
    except FileNotFoundError:
        print(f"File not found: {file_path}")
    except Exception as e:
        print(f"Error: {e}")
